from src.logic.module import (
    search_path,
    find_path_by_partial_name,
    get_column_from_csv,
)
from src.logic.before_packing import (
    get_adjusted_unit_weight,
    convert_to_cafe24_product_code,
    merge_product_instructions,
    report_missing_instructions,
    determine_box_size,
    flatten_order_items_by_order_number,
)
import pandas as pd
import os
import webbrowser
from datetime import datetime  # 폴더이름에 현재 날짜 넣기 위해

from pathlib import Path
import sys

sys.path.append(str(Path(__file__).resolve().parent.parent.parent))
from settings.column_mapping import KOR_TO_ENG_COLUMN_MAP, ENG_TO_KOR_COLUMN_MAP

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.page import PageMargins

import logging

logger = logging.getLogger(__name__)


def prepare_to_pack(log_set_callback, log_get_callback):
    try:
        log_set_callback("🐶🐶🐶시작! 프로그램 실행🐶🐶🐶")

        ### Output folder
        output_folder = "result_" + datetime.now().strftime(
            "%a.%H.%M.%S"
        )  # 요일.시.분.초
        os.makedirs(output_folder)
        logger.info("Output folder created")

        ########################################## Split file from download folder ##########################################
        # Get the folder path from setting\path.csv where the raw file is located
        download_from_internet_path = search_path()

        # Find the file downloaded from Cafe24
        # File name format downloaded from Cafe24: lalapetmall_today's_date_serialnumber_serialnumber
        download_from_cafe24_path = find_path_by_partial_name(
            download_from_internet_path,
            "lalapetmall_" + datetime.today().strftime("%Y%m%d") + "_",
        )
        raw_data_df = pd.read_csv(download_from_cafe24_path)

        # log
        log_set_callback("다운로드 받은 파일 검색")
        logger.info("Find the file downloaded from Cafe24")

        ### Split into three files
        # Order list file
        order_list_path = rf"{output_folder}\order_list.xlsx"
        order_list_header_list = get_column_from_csv(
            r"settings\header.csv", "order_list_header"
        )
        logger.info("Get columns from settings\\header.csv")
        order_list_df = raw_data_df[order_list_header_list].rename(
            columns=KOR_TO_ENG_COLUMN_MAP
        )
        logger.info("Split into three files: 1. order list")

        # Hanjin file
        hanjin_path = rf"{output_folder}\upload_to_hanjin.xlsx"
        hanjin_header_list = get_column_from_csv(
            r"settings\header.csv", "hanjin_header"
        )
        hanjin_list_df = raw_data_df[hanjin_header_list].rename(
            columns=KOR_TO_ENG_COLUMN_MAP
        )

        # Replace product_name with internal_product_name if not NaN
        hanjin_list_df.loc[
            hanjin_list_df["product_name_with_option"].str.contains("개인", na=False),
            "product_name_with_option",
        ] = hanjin_list_df["internal_product_name"]
        hanjin_list_df = hanjin_list_df.drop(columns=["internal_product_name"])

        logger.info("Split into three files: 2. hanjin")

        # Cafe24 upload file
        cafe24_upload_path = rf"{output_folder}\excel_sample_old.csv"
        cafe24_upload_header_list = get_column_from_csv(
            r"settings\header.csv", "cafe24_upload"
        )
        df_cafe24_upload = raw_data_df[cafe24_upload_header_list]
        df_cafe24_upload.to_csv(
            cafe24_upload_path,
            index=False,
            encoding="utf-8-sig",
        )
        logger.info("Split into three files: 3. cafe24")

        # log
        log_set_callback("헤더명에 따라 세 개의 파일로 분리")

        ########################################## Print out product instruction ##########################################
        converted_cafe24_codes, missing_codes = convert_to_cafe24_product_code(
            order_list_df
        )
        order_list_df["product_code"] = converted_cafe24_codes
        logger.info("converted_cafe24_codes")

        if missing_codes:
            logger.warning("Unmatched product codes: %s", missing_codes)

        not_found_files = merge_product_instructions(
            output_folder, converted_cafe24_codes
        )
        logger.info("product instructions merged")
        report_missing_instructions(output_folder, order_list_df, not_found_files)

        # log
        log_set_callback("상품 설명지 병합")
        logger.info("Merge product instructions")

        ########################################## Data Transformation(Determine box size) ##########################################
        #### Adjust weight data
        order_list_df["unit_weight"] = order_list_df.apply(
            get_adjusted_unit_weight, axis=1
        )

        order_list_df["item_total_weight"] = (
            order_list_df["unit_weight"] * order_list_df["quantity"]
        )

        total_weight_by_order = order_list_df.groupby("order_number")[
            "item_total_weight"
        ].sum()
        order_list_df["total_weight_by_order"] = order_list_df["order_number"].map(
            total_weight_by_order
        )
        logger.info("Weight data adjusted")

        ### Determine box size
        order_list_df["box_size"] = order_list_df.apply(determine_box_size, axis=1)
        logger.info(
            order_list_df[
                ["order_number", "product_name", "total_weight_by_order", "box_size"]
            ]
        )

        # log
        logger.info("Determine box size")
        log_set_callback("주문리스트의 박스 정보 입력")

        ########################################## Data transformation using pandas ##########################################
        ### Generate serial numbers for unique order_numbers; set blank for subsequent duplicates
        # Step 1: Mark True for the first occurrence of each order_number
        order_list_df["serial_number"] = ~order_list_df["order_number"].duplicated()

        # Step 2: Assign a running number to the first occurrences. cumsum: cumulative sum
        order_list_df["serial_number"] = order_list_df["serial_number"].cumsum()

        logger.info("Serial numbers generated")

        # log
        log_set_callback("주문리스트의 일련번호 입력")

        ###
        # Columns to clean duplicates
        cols_to_clean = [
            "order_number",
            "orderer_name",
            "recipient_name",
            "recipient_address",
            "delivery_message",
            "box_size",
        ]

        # For each column, replace duplicate values with ''
        for col in cols_to_clean:
            order_list_df[col] = order_list_df[col].where(
                ~order_list_df["serial_number"].duplicated(), ""
            )
        logger.info("Clean duplicated value")

        # If 'pickup' column contains '방문수령', replace it with 'O'
        order_list_df.loc[order_list_df["pickup"] == "방문수령", "pickup"] = "O"
        logger.info("Reorganize pickup column")

        # Replace product_name with internal_product_name if not NaN
        order_list_df.loc[
            order_list_df["product_name"].str.contains("개인", na=False),
            "product_name",
        ] = order_list_df["internal_product_name"]

        ### Reorder column
        column_order = [
            "serial_number",
            "order_number",
            "orderer_name",
            "product_name",
            "option",
            "quantity",
            "recipient_name",
            "price",
            "recipient_address",
            "delivery_message",
            "box_size",
            "pickup",
            "subscription_cycle",
            "membership_level",
        ]
        order_list_df[column_order].to_excel(order_list_path, index=False)
        ########################################## Document design for order list(fill the color) using openpyxl ##########################################

        order_list_df = pd.read_excel(order_list_path)
        # Load the workbook and worksheet
        wb = load_workbook(order_list_path)
        ws = wb.active
        max_row = ws.max_row

        # Define color fills
        red_fill = PatternFill(
            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
        )  # RGB(255,204,207)
        gray_fill = PatternFill(
            start_color="BEBEBE", end_color="BEBEBE", fill_type="solid"
        )  # RGB(190,190,190)
        blue_fill = PatternFill(
            start_color="C0E6F5", end_color="C0E6F5", fill_type="solid"
        )  # RGB(192,230,245)

        ### Apply conditional formatting
        # quantity >= 2 → Red
        quantity_col = order_list_df.columns.get_loc("quantity") + 1
        logger.info("quantity_col: %s", quantity_col)
        ws.conditional_formatting.add(
            f"{chr(64 + quantity_col)}2:{chr(64 + quantity_col)}{max_row}",
            Rule(
                type="expression",
                dxf=DifferentialStyle(fill=red_fill),
                formula=[f"${chr(64 + quantity_col)}2>=2"],
            ),
        )

        # subscription_cycle not blank → Blue
        logger.info("df_order_list: %s", order_list_df)
        product_col = order_list_df.columns.get_loc("product_name") + 1
        sub_col = order_list_df.columns.get_loc("subscription_cycle") + 1
        logger.info("sub_col: %s", sub_col)

        product_col_letter = chr(64 + product_col)
        sub_col_letter = chr(64 + sub_col)

        ws.conditional_formatting.add(
            f"{product_col_letter}2:{product_col_letter}{max_row}",
            Rule(
                type="expression",
                dxf=DifferentialStyle(fill=blue_fill),
                formula=[f"LEN(TRIM(${sub_col_letter}2))>0"],
            ),
        )

        # order_number duplicates → Gray
        serial_num_col = order_list_df.columns.get_loc("serial_number") + 1
        serial_numbers = order_list_df["serial_number"]
        duplicates = serial_numbers[serial_numbers.duplicated(keep=False)].unique()

        for row in range(2, max_row + 1):
            cell_value = ws.cell(row=row, column=serial_num_col).value
            if cell_value in duplicates:
                ws.cell(row=row, column=serial_num_col).fill = gray_fill
                ws.cell(row=row, column=serial_num_col + 1).fill = gray_fill

        ########################################## Document design for order list ##########################################
        price_col = order_list_df.columns.get_loc("price") + 1
        for row in ws.iter_rows(min_row=2, min_col=price_col, max_col=price_col):
            for cell in row:
                cell.number_format = "#,##0"

        ### Border
        # Define border styles
        thin_side = Side(style="thin")  # Like xlThin in VBA

        # inside horizontal borders
        for row_idx in range(2, max_row):  # Exclude last row
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = Border(bottom=thin_side)

        ### Insert the box size counts at the buttom
        # Count the occurrences of each box size and convert the result into a DataFrame
        box_size_counts = order_list_df["box_size"].value_counts().reset_index()
        box_size_counts.columns = ["박스", "개수"]

        custom_order = [1, 2, 3, 420, 287]

        # Convert '박스' column to categorical with custom order
        box_size_counts["박스"] = pd.Categorical(
            box_size_counts["박스"], categories=custom_order, ordered=True
        )

        # Sort according to custom order
        box_size_counts = box_size_counts.sort_values("박스").reset_index(drop=True)

        start_row = max_row + 3

        # Write headers from DataFrame
        for col_idx, col_name in enumerate(box_size_counts.columns, start=1):
            ws.cell(row=start_row, column=col_idx + 1, value=col_name)

        # Write data rows
        for row_idx, row in enumerate(box_size_counts.itertuples(index=False), start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=start_row + row_idx, column=col_idx + 1, value=value)

        ### Set column width
        # Mapping of column headers to desired widths
        column_widths = {
            "serial_number": 3,
            "order_number": 9,
            "orderer_name": 7,
            "product_name": 38,
            "option": 16,
            "quantity": 4,
            "recipient_name": 7,
            "price": 8,
            "recipient_address": 28,
            "delivery_message": 14,
            "box_size": 5,
            "pickup": 3,
        }

        for col in order_list_df.columns.to_list():
            col_idx = order_list_df.columns.get_loc(col) + 1
            col_letter = chr(64 + col_idx)
            if col in column_widths.keys():
                ws.column_dimensions[col_letter].width = column_widths.get(col)
            else:
                ws.column_dimensions[col_letter].hidden = True

        ### Rename header row
        for cell in ws[1]:  # First row
            eng_col = cell.value
            if eng_col in ENG_TO_KOR_COLUMN_MAP:
                cell.value = ENG_TO_KOR_COLUMN_MAP[eng_col]

        ### Apply wrap text to all cells in the data range
        for row in ws.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        ### Print settings
        # Set header rows to repeat when printing
        ws.print_title_rows = "$1:$1"  # Repeat row 1 on each printed page

        # Set headers and footers
        ws.oddHeader.left.text = "&D &T"  # Date and time
        ws.oddHeader.center.text = "전채널 주문 리스트"
        ws.oddHeader.right.text = "&P/&N"  # Page X of N

        # Set margins (in inches)
        ws.page_margins = PageMargins(
            left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3
        )

        ## Set paper size to A4
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # Set page orientation to landscape
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

        ### Save styled workbook
        wb.save(order_list_path)

        ########################################## Multiple items process for hanjin list ##########################################
        flatten_order_items_by_order_number(hanjin_list_df).rename(
            columns=ENG_TO_KOR_COLUMN_MAP
        ).to_excel(hanjin_path, index=False)

        # log
        log_set_callback("한진 사이트에 업로드할 파일 작성")

        ########################################## ##########################################

        ####한진택배 사이트 열기
        webbrowser.open("https://focus.hanjin.com/login")

        ####result 폴더 열기
        os.startfile(f"{output_folder}")
        # log
        log_set_callback("🐶🐶🐶끝! 실행 완료🐶🐶🐶")

    except Exception as e:
        log_set_callback(f"⚠️오류 발생: {e}")
